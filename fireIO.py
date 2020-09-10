# 该程序可批量发放邮件。可通过更改文件夹下.xlsx文件内容给相应邮箱发放相应内容。该程序可方便老师/助教通过邮箱发布学生学科成绩
# 不同发件人，不同接收者的情况下，除了更改文件夹下.xlsx文件之外，同时需要改动本程序内部部分：
# 1）请将发件人邮箱账号和发件人序列号做出相应更改。发件人可自行访问邮箱网页版获得该序列号（激活码）
# 2）请在明显标注需要修改的地方，将‘max_row’的值改为n+1，n为学生总数
# 3）请在程序最后将学生总数改至提示的地方
import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.utils import formataddr


my_sender = ''  # 发件人邮箱账号

my_user = []
result = []
namelist = []
contents = []

wb = openpyxl.load_workbook('try-xlsxIO.xlsx')
ws = wb.active

stu_num = ws.max_row

 #get users' adress
for each_row in ws.iter_rows(min_row=2, min_col=2, max_row=stu_num, max_col=2):
    my_user.append(each_row[0].value+"@shanghaitech.edu.cn")

#get students' name
for each_row in ws.iter_rows(min_row=2, min_col=1, max_row=stu_num, max_col=1):
    namelist.append(str(each_row[0].value))

#content input
for i in range(0,stu_num-1):
    contents.append(
"""{name}
""".format(name=namelist[i]))


def mail(k, contents):
    ret = True
    try:
        msg = MIMEText(contents, 'plain', 'utf-8')
        msg['From'] = formataddr(["", my_sender])
        msg['Subject'] = ""

        server = smtplib.SMTP("smtp.shanghaitech.edu.cn", 25)
        server.set_debuglevel(1)
        server.sendmail(my_sender, [my_user[k], ], msg.as_string())
        server.quit()
       
    except Exception:
        ret = False
    return ret
######################请将（0，2）处的2更改为学生总数##############################
for i in range(0, stu_num-1):
    ret = mail(i, contents[i])
    if ret == True:
        print("邮件发送成功")
    else:
        print(str(i)+"号邮件发送失败")

